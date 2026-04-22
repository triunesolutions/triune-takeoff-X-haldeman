import io
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .constants import TRIUNE_COLUMNS


def _fg(h: str) -> str:
    return (h or "#FFFFFF").replace("#", "").upper()


def style_takeoff_and_raw_bytes(
    source_df: pd.DataFrame,
    takeoff_df: pd.DataFrame,
    header_hex: str,
    product_hex: str,
    zebra_hex: str,
    qty_gold_hex: str,
    bold_cols: List[str],
) -> bytes:
    wb = Workbook()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor=_fg(header_hex))
    product_fill = PatternFill("solid", fgColor=_fg(product_hex))
    zebra_fill = PatternFill("solid", fgColor=_fg(zebra_hex))
    qty_fill = PatternFill("solid", fgColor=_fg(qty_gold_hex))

    bold_set = set(bold_cols or [])

    ws = wb.active
    ws.title = "Triune Takeoff Haldeman"

    for c_idx, col in enumerate(TRIUNE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="000000")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, row in enumerate(takeoff_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            col_name = TRIUNE_COLUMNS[c_idx - 1]
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.font = Font(bold=(col_name in bold_set), color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_name == "QTY":
                cell.number_format = "#,##0"

    qty_idx = TRIUNE_COLUMNS.index("QTY") + 1
    prod_idx = 1

    for r in range(2, ws.max_row + 1):
        v_prod = ws.cell(r, prod_idx).value or ""
        row_vals = [ws.cell(r, c).value for c in range(1, len(TRIUNE_COLUMNS) + 1)]
        is_blank_row = all((v is None or v == "") for v in row_vals)
        is_pt = isinstance(v_prod, str) and v_prod.endswith(" Total") and v_prod != "Grand Total"
        is_gt = v_prod == "Grand Total"

        if is_blank_row:
            continue
        if is_pt:
            for c in range(1, len(TRIUNE_COLUMNS) + 1):
                cc = ws.cell(r, c)
                cc.fill = product_fill
                cc.font = Font(bold=True, color="000000")
                cc.border = border
        elif is_gt:
            for c in range(1, len(TRIUNE_COLUMNS) + 1):
                cc = ws.cell(r, c)
                cc.fill = header_fill
                cc.font = Font(bold=True, color="000000")
                cc.border = border
        else:
            if r % 2 == 0:
                for c in range(1, len(TRIUNE_COLUMNS) + 1):
                    ws.cell(r, c).fill = zebra_fill
            ws.cell(r, qty_idx).fill = qty_fill

    for i in range(1, len(TRIUNE_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "A2"
    try:
        ws.sheet_view.zoomScale = 80
    except Exception:
        pass

    ws_raw = wb.create_sheet(title="RawData")
    for c_idx, col in enumerate(source_df.columns, start=1):
        cell = ws_raw.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="000000")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, row in enumerate(source_df.itertuples(index=False, name=None), start=2):
        for c_idx, value in enumerate(row, start=1):
            colname = source_df.columns[c_idx - 1]
            if colname.upper() == "QTY":
                try:
                    s = str(value).replace(",", "").strip()
                    if s:
                        num = float(s)
                        value = int(num) if num.is_integer() else num
                except Exception:
                    pass
            cell = ws_raw.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

    for i, col in enumerate(source_df.columns, 1):
        try:
            max_len = max(source_df[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws_raw.column_dimensions[get_column_letter(i)].width = max_len

    try:
        ws_raw.sheet_view.zoomScale = 80
    except Exception:
        pass

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
