"""Convert AirDistribution_CrossRef.xlsx → api/_lib/crossref.json.

Run once after the Excel source updates:
    python scripts/build-crossref.py

Reads the 'Master Cross-Reference' sheet. Each data row has a Category
plus one or more brand columns filled with that brand's equivalent model(s).
Writes a compact JSON with brands, categories, brands_by_category, and data rows.
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.environ.get(
    "CROSSREF_XLSX",
    r"C:\Users\TriuneTakeoff\Downloads\Cross Reference Tool files\AirDistribution_CrossRef.xlsx",
)
OUT = os.path.join(ROOT, "api", "_lib", "crossref.json")
MASTER_SHEET = "📋 Master Cross-Reference"


def main() -> int:
    if not os.path.exists(XLSX):
        print(f"[build-crossref] source not found: {XLSX}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        print(f"[build-crossref] '{MASTER_SHEET}' sheet missing", file=sys.stderr)
        return 1

    ws = wb[MASTER_SHEET]
    rows = ws.iter_rows(values_only=True)

    # Row 0 = title, row 1 = header
    next(rows)
    header = list(next(rows))
    # Columns: [Dataset, Category, Description, <brand 1>, <brand 2>, ...]
    if header[:3] != ["Dataset", "Category", "Description"]:
        print(f"[build-crossref] unexpected header row: {header[:5]}", file=sys.stderr)
        return 1

    brands = [h for h in header[3:] if h]
    brand_cols = [i + 3 for i, h in enumerate(header[3:]) if h]

    data: list = []
    categories_seen: list = []
    categories_set: set = set()
    brands_by_cat: dict = {}

    for row in rows:
        if not row:
            continue
        category = (row[1] or "").strip() if row[1] else ""
        description = (row[2] or "").strip() if row[2] else ""
        if not category:
            continue

        brand_values: dict = {}
        for brand, col in zip(brands, brand_cols):
            v = row[col] if col < len(row) else None
            if v is None:
                continue
            s = str(v).strip()
            if s:
                brand_values[brand] = s
        if not brand_values:
            continue

        if category not in categories_set:
            categories_set.add(category)
            categories_seen.append(category)

        bset = brands_by_cat.setdefault(category, set())
        for b in brand_values:
            bset.add(b)

        data.append({
            "category": category,
            "description": description,
            "brands": brand_values,
        })

    # Sort brands_by_category to match the original brand order from the header
    brand_order = {b: i for i, b in enumerate(brands)}
    brands_by_category = {
        c: sorted(list(bs), key=lambda b: brand_order.get(b, 999))
        for c, bs in brands_by_cat.items()
    }

    out = {
        "brands": brands,
        "categories": categories_seen,
        "brands_by_category": brands_by_category,
        "data": data,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    print(f"[build-crossref] wrote {OUT}")
    print(f"  brands: {len(brands)}")
    print(f"  categories: {len(categories_seen)}")
    print(f"  data rows: {len(data)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
